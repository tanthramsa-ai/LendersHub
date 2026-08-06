#!/usr/bin/env python3
"""
Watches the shared QA defect spreadsheet and reports what changed since the last run.

The sheet is exported as a whole workbook (not per-tab CSV) so new and removed tabs
are caught too -- the "New Req" tab appeared that way and would have been invisible
to a single-tab CSV poll.

Screenshots pasted into the sheet do not survive as text, so the per-tab image count
is tracked separately: a screenshot-only edit still shows up as a change.

Usage:
    python scripts/sheet-watch.py            # report changes since last run, then save
    python scripts/sheet-watch.py --init     # seed the snapshot without reporting
    python scripts/sheet-watch.py --dry-run  # report changes but do NOT save
    python scripts/sheet-watch.py --json     # machine-readable output

Exit codes:
    0  ran fine (whether or not anything changed)
    2  fetch or parse failed
"""

from __future__ import annotations

import argparse
import io
import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(2)

SHEET_ID = os.environ.get(
    "QA_SHEET_ID", "1DOBd7i_gHHAzomvUrtY8d41EkTmA1zqM-9jPmS7vOm0"
)
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

HERE = os.path.dirname(os.path.abspath(__file__))
SNAPSHOT_PATH = os.path.join(HERE, ".sheet-watch", "snapshot.json")

# Rows that are structural noise rather than defects -- they carry no information on
# their own and would otherwise churn the diff whenever a screenshot moves.
NOISE = {"important", "before", "after", "expected", "actual", "suggestion"}


def fetch_workbook(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "lendershub-sheet-watch/1"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def normalise(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def extract(data: bytes) -> dict:
    """Reduce the workbook to {tab: {"rows": [...], "images": n}}."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    tabs = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for r in range(1, ws.max_row + 1):
            cells = [normalise(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            cells = [c for c in cells if c]
            if not cells:
                continue
            line = " | ".join(cells)
            if line.strip().lower().rstrip(":") in NOISE:
                continue
            rows.append(line)
        tabs[name] = {"rows": rows, "images": len(ws._images)}
    return tabs


def load_snapshot(path: str) -> dict | None:
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def save_snapshot(path: str, tabs: dict) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    payload = {
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "sheetId": SHEET_ID,
        "tabs": tabs,
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


def diff(old_tabs: dict, new_tabs: dict) -> dict:
    result = {
        "tabsAdded": sorted(set(new_tabs) - set(old_tabs)),
        "tabsRemoved": sorted(set(old_tabs) - set(new_tabs)),
        "changed": {},
    }
    for name in sorted(set(new_tabs) & set(old_tabs)):
        old, new = old_tabs[name], new_tabs[name]
        # Multiset-free comparison: duplicate lines are rare here and treating rows as
        # a set keeps a row moving up or down the sheet from reading as a change.
        added = [r for r in new["rows"] if r not in set(old["rows"])]
        removed = [r for r in old["rows"] if r not in set(new["rows"])]
        img_delta = new["images"] - old["images"]
        if added or removed or img_delta:
            result["changed"][name] = {
                "added": added,
                "removed": removed,
                "imageDelta": img_delta,
            }
    # A brand-new tab's rows are all new content worth showing.
    for name in result["tabsAdded"]:
        result["changed"][name] = {
            "added": new_tabs[name]["rows"],
            "removed": [],
            "imageDelta": new_tabs[name]["images"],
        }
    return result


def has_changes(d: dict) -> bool:
    return bool(d["tabsAdded"] or d["tabsRemoved"] or d["changed"])


def render(d: dict, new_tabs: dict) -> str:
    if not has_changes(d):
        return "NO CHANGES"

    out = ["SHEET CHANGED", ""]
    if d["tabsAdded"]:
        out.append(f"New tabs: {', '.join(d['tabsAdded'])}")
    if d["tabsRemoved"]:
        out.append(f"Removed tabs: {', '.join(d['tabsRemoved'])}")
    if d["tabsAdded"] or d["tabsRemoved"]:
        out.append("")

    for name, ch in d["changed"].items():
        out.append(f"--- {name} ---")
        for row in ch["added"]:
            out.append(f"  + {row}")
        for row in ch["removed"]:
            out.append(f"  - {row}")
        if ch["imageDelta"]:
            sign = "+" if ch["imageDelta"] > 0 else ""
            out.append(f"  ({sign}{ch['imageDelta']} screenshot(s))")
        out.append("")

    total = sum(len(t["rows"]) for t in new_tabs.values())
    out.append(f"({total} rows across {len(new_tabs)} tabs)")
    return "\n".join(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="Report changes to the QA defect sheet.")
    ap.add_argument("--init", action="store_true", help="seed the snapshot, report nothing")
    ap.add_argument("--dry-run", action="store_true", help="report but do not save the snapshot")
    ap.add_argument("--json", action="store_true", help="emit JSON instead of text")
    args = ap.parse_args()

    try:
        data = fetch_workbook(EXPORT_URL)
    except (urllib.error.URLError, urllib.error.HTTPError) as exc:
        sys.stderr.write(f"Failed to fetch the sheet: {exc}\n")
        sys.stderr.write("If this started failing, link sharing may have been turned off.\n")
        return 2

    try:
        new_tabs = extract(data)
    except Exception as exc:  # noqa: BLE001 -- surface any parse failure verbatim
        sys.stderr.write(f"Failed to parse the workbook: {exc}\n")
        return 2

    previous = load_snapshot(SNAPSHOT_PATH)

    if args.init or previous is None:
        save_snapshot(SNAPSHOT_PATH, new_tabs)
        total = sum(len(t["rows"]) for t in new_tabs.values())
        msg = f"Snapshot seeded: {total} rows across {len(new_tabs)} tabs ({', '.join(new_tabs)})"
        print(json.dumps({"seeded": True, "tabs": list(new_tabs)}) if args.json else msg)
        return 0

    d = diff(previous["tabs"], new_tabs)

    if args.json:
        print(json.dumps({"changed": has_changes(d), "since": previous["fetchedAt"], **d}, indent=2, ensure_ascii=False))
    else:
        print(render(d, new_tabs))

    if not args.dry_run:
        save_snapshot(SNAPSHOT_PATH, new_tabs)

    return 0


if __name__ == "__main__":
    sys.exit(main())
